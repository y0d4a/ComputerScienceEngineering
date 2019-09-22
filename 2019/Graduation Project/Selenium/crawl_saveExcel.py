import datetime
import openpyxl
import re
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

# 엑셀 파일 생성 함수
def make_excel(user_name, keyword):

    # 현재 날짜
    now = datetime.datetime.now()
    date = now.strftime('%Y.%m.%d')
    # 엑셀 파일 저장 위치
    excel_file_path = 'C:/Users/battl/PycharmProjects/myProject02/excel/'
    # 엑셀 파일 이름
    excel_file_name = excel_file_path + user_name + '_' + keyword + '_' + date + '.xlsx'
    # 엑셀 sheet 이름
    excel_sheet_title = 'confirm'

    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    # 헤더 입력
    sheet1.cell(row=1, column=1).value = 'user_name'
    sheet1.cell(row=1, column=2).value = 'user_mail'
    sheet1.cell(row=1, column=3).value = 'curr_url'
    sheet1.cell(row=1, column=4).value = 'prev_url'
    sheet1.cell(row=1, column=5).value = 'pageList'
    sheet1.cell(row=1, column=6).value = 'relativeKeywordList'
    sheet1.cell(row=1, column=7).value = 'level'
    sheet1.cell(row=1, column=8).value = 'paths'
    sheet1.cell(row=1, column=9).value = 'keyword'
    sheet1.cell(row=1, column=10).value = 'sub_keyword'
    sheet1.cell(row=1, column=11).value = 'pageContents'
    sheet1.cell(row=1, column=12).value = 'memo'
    sheet1.cell(row=1, column=13).value = 'tags'
    sheet1.cell(row=1, column=14).value = 'nowTime'
    sheet1.cell(row=1, column=15).value = 'screenshot'

    work_book.save(filename=excel_file_name)
    work_book.close()

# 엑셀 셀 너비 자동 조정 함수
def adjust_column_width(ws):

    for col in ws.columns:
        max_length = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjust_width = max_length + 2
        ws.column_dimensions[col_name].width = adjust_width

# 엑셀에 크롤링된 데이터 저장 함수
def crawl_saveExcel(user_name, user_email, curr_url, prev_url, pageList, relativeKeywordList, level, paths, keyword, sub_keyword, pageContents, memo, screenshot, tagged, nowTime):

    # 현재 날짜
    now = datetime.datetime.now()
    date = now.strftime('%Y.%m.%d')
    # 엑셀 파일 저장 위치
    excel_file_path = 'C:/Users/battl/PycharmProjects/myProject02/excel/'
    # 엑셀 파일 이름
    excel_file_name = excel_file_path + user_name + '_' + keyword + '_' + date + '.xlsx'
    # 엑셀 sheet 이름
    excel_sheet_title = 'confirm'

    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 1

    crawl_info = {
        "user_name": user_name,
        "user_email": user_email,
        "curr_url": curr_url,
        "prev_url": prev_url,
        "pageList": pageList,
        "relativeKeywordList": relativeKeywordList,
        "level": int(level),
        "paths": paths,
        "keyword": keyword,
        "sub_keyword": sub_keyword,
        "pageContents": pageContents,
        "memo": memo,
        "tagged": tagged,
        "nowTime": nowTime,
        "screenshot": screenshot,
    }

    crawl_info_values = list(crawl_info.values())

    for idx, info_values in enumerate(crawl_info_values):
        # 페이지 리스트(4), 서브 키워드(9), 본문 요약(10)인 경우(리스트 형태)
        # 페이지 리스트(4)와 서브 키워드(9), 본문 요약(10)의 경우는 검색 페이지와 홈페이지인 경우 추출이 안될 수 있기 때문에
        # 빈 리스트를 고려해주어야 한다.
        if idx == 4 or idx == 9 or idx == 10:
            if info_values == []:
                sheet1.cell(row=excel_row, column=excel_column).value = ''
            else:
                for info in info_values:
                    sheet1.cell(row=excel_row, column=excel_column).value = info
                    excel_row += 1

        # 연관 검색어 URL 리스트인 경우(딕셔너리 형태)(5)
        # 연관 검색어 URL 리스트(5) 역시 검색 페이지에서만 추출되기 때문에 홈페이지인 경우를 고려해 빈 리스트를 고려해주어야 한다.
        elif idx == 5:
            if info_values == []:
                sheet1.cell(row=excel_row, column=excel_column).value = ''
            else:
                relativeKey = []
                for i in info_values.keys():
                    relativeKey.append(i + "(" + info_values[i] + ")")

                for r in relativeKey:
                    sheet1.cell(row=excel_row, column=excel_column).value = r
                    excel_row += 1

        # 레벨의 경우(상수 형태)
        elif idx == 6:
            sheet1.cell(row=excel_row, column=excel_column).value = info_values

        # 스크린 샷(14)인 경우(문자열 형태)
        elif idx == 14:
            png_loc = 'C:/Users/battl/PycharmProjects/myProject02/screenshot/' + user_name + '_' + keyword + '.png'
            png = openpyxl.drawing.image.Image(png_loc)
            sheet1.add_image(png, 'P2')

        # 사용자명, 이메일, 현재 페이지, 이전 페이지, 경로, 키워드, 메모, 태그, 추출 시간인 경우(문자열 형태)
        else:
            if info_values == []:
                sheet1.cell(row=excel_row, column=excel_column).value = ''
            else:
                sheet1.cell(row=excel_row, column=excel_column).value = info_values

        # 다음 인자를 넣기 위해 셀의 열 이동
        excel_row = 2
        excel_column += 1

    # 엑셀 파일 크기 조정하기
    adjust_column_width(sheet1)

    # 엑셀 파일 저장하기
    excel_file.save(excel_file_name)
    excel_file.close()